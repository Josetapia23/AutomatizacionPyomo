"""
Módulo OPTIMIZADO CORREGIDO para el procesamiento de ofertas.
Versión que mantiene la velocidad pero corrige los cálculos para precisión exacta.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import logging
import re
from pathlib import Path
import time
from functools import lru_cache
import openpyxl
from openpyxl import load_workbook

from config import DATOS_INICIALES, OFERTAS_DIR, RESULTADO_OFERTAS
from core.validador_ofertas import ValidadorNombresOfertas
from core.utils import (
    verificar_archivo_existe,
    verificar_hoja_existe,
    leer_excel_seguro,
    guardar_excel_seguro,
    solicitar_input_seguro,
    fecha_a_texto,
    leer_excel_case_insensitive,
    verificar_hoja_existe_case_insensitive
)
from core.indexadores import calcular_numerador, calcular_denominador, crear_proyeccion_precio_sicep

logger = logging.getLogger(__name__)

# ============================================================================
# MANTENER FUNCIONES DE LECTURA RÁPIDA (ESTAS ESTÁN BIEN)
# ============================================================================

def leer_excel_ultra_rapido(archivo_path, hojas_necesarias):
    """
    Lee múltiples hojas de Excel de una vez usando openpyxl optimizado.
    VERSIÓN CORREGIDA: Ahora con soporte case-insensitive.
    Hasta 3x más rápido que pd.read_excel iterativo.
    """
    start_time = time.time()
    resultados = {}
    
    try:
        workbook = load_workbook(archivo_path, read_only=True, data_only=True)
        hojas_disponibles = workbook.sheetnames
        
        # NUEVO: Crear mapeo case-insensitive
        mapeo_hojas = {}
        for hoja_objetivo in hojas_necesarias:
            hoja_objetivo_lower = hoja_objetivo.lower()
            hoja_encontrada = None
            
            for hoja_disponible in hojas_disponibles:
                if hoja_disponible.lower() == hoja_objetivo_lower:
                    hoja_encontrada = hoja_disponible
                    break
            
            if hoja_encontrada:
                mapeo_hojas[hoja_objetivo] = hoja_encontrada
                if hoja_encontrada != hoja_objetivo:
                    print(f"✅ Mapeo case-insensitive: '{hoja_objetivo}' → '{hoja_encontrada}'")
            else:
                logger.warning(f"Hoja '{hoja_objetivo}' no encontrada en {archivo_path}")
                logger.warning(f"Hojas disponibles: {hojas_disponibles}")
                resultados[hoja_objetivo] = pd.DataFrame()
                continue
        
        # Leer cada hoja mapeada
        for hoja_objetivo, hoja_real in mapeo_hojas.items():
            try:
                worksheet = workbook[hoja_real]
                
                data_rows = []
                headers = None
                
                for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if not any(cell is not None and str(cell).strip() != '' for cell in row):
                        continue
                        
                    if headers is None:
                        headers = [str(cell) if cell is not None else f"COL_{i}" for i, cell in enumerate(row)]
                    else:
                        data_rows.append(row)
                
                if headers and data_rows:
                    data_truncated = []
                    for row in data_rows:
                        if len(row) > len(headers):
                            data_truncated.append(row[:len(headers)])
                        else:
                            row_padded = list(row) + [None] * (len(headers) - len(row))
                            data_truncated.append(row_padded)
                    
                    resultados[hoja_objetivo] = pd.DataFrame(data_truncated, columns=headers)
                else:
                    resultados[hoja_objetivo] = pd.DataFrame()
                    
            except Exception as e:
                logger.error(f"Error al leer hoja '{hoja_real}': {e}")
                resultados[hoja_objetivo] = pd.DataFrame()
        
        workbook.close()
        
        end_time = time.time()
        tiempo_lectura = end_time - start_time
        logger.debug(f"Lectura rápida case-insensitive de {len(hojas_necesarias)} hojas: {tiempo_lectura:.2f}s")
        
        return resultados
        
    except Exception as e:
        logger.error(f"Error en lectura rápida case-insensitive de {archivo_path}: {e}")
        return leer_excel_fallback(archivo_path, hojas_necesarias)

def leer_excel_fallback(archivo_path, hojas_necesarias):
    """
    Método de respaldo usando pandas tradicional.
    VERSIÓN CORREGIDA: Ahora con soporte case-insensitive.
    """
    resultados = {}
    
    try:
        excel_file = pd.ExcelFile(archivo_path)
        hojas_disponibles = excel_file.sheet_names
        
        for hoja_objetivo in hojas_necesarias:
            hoja_objetivo_lower = hoja_objetivo.lower()
            hoja_encontrada = None
            
            # Buscar coincidencia case-insensitive
            for hoja_disponible in hojas_disponibles:
                if hoja_disponible.lower() == hoja_objetivo_lower:
                    hoja_encontrada = hoja_disponible
                    break
            
            if hoja_encontrada:
                resultados[hoja_objetivo] = pd.read_excel(excel_file, sheet_name=hoja_encontrada)
                if hoja_encontrada != hoja_objetivo:
                    print(f"✅ Fallback case-insensitive: '{hoja_objetivo}' → '{hoja_encontrada}'")
            else:
                logger.warning(f"Hoja '{hoja_objetivo}' no encontrada en fallback")
                resultados[hoja_objetivo] = pd.DataFrame()
                
    except Exception as e:
        logger.error(f"Error en fallback case-insensitive: {e}")
        # Último recurso: DataFrames vacíos
        for hoja in hojas_necesarias:
            resultados[hoja] = pd.DataFrame()
    
    return resultados
# ============================================================================
# CORREGIR CÁLCULO DE INDEXADORES - USAR LÓGICA ORIGINAL
# ============================================================================

def precomputar_indexadores_corregido(indexadores_df, proyeccion_df):
    """
    Pre-calcula indexadores pero mantiene la estructura original para precisión.
    """
    print("🔧 Pre-calculando indexadores (versión corregida)...")
    start_time = time.time()
    
    # Combinar datos históricos y proyectados
    indexadores_completo = pd.concat([indexadores_df, proyeccion_df], ignore_index=True)
    indexadores_completo = indexadores_completo.drop_duplicates(subset=['fechaoperacion'])
    
    print(f"✅ Indexadores pre-calculados: {len(indexadores_completo)} fechas únicas en {time.time() - start_time:.2f}s")
    
    return indexadores_completo

def calcular_numerador_rapido(fecha, indexador, numerador, indexadores_completo_df):
    """
    Versión optimizada del cálculo de numerador usando DataFrame pre-cargado.
    MANTIENE la lógica exacta de la función original.
    """
    fecha_ano_mes = fecha_a_texto(fecha)
    
    # Buscar en el DataFrame combinado (más rápido que dos búsquedas)
    mask = indexadores_completo_df['fechaoperacion'].apply(lambda x: fecha_a_texto(x)) == fecha_ano_mes
    matching_rows = indexadores_completo_df[mask]
    
    if matching_rows.empty:
        return None
    
    row = matching_rows.iloc[0]
    
    # LÓGICA EXACTA de la función original
    if indexador == "IPC":
        return row['ipc']
    elif indexador != "IPC" and numerador == "PROVISIONAL":
        return row['oferta_interna_prov']
    elif indexador != "IPC" and numerador == "DEFINITIVO":
        return row['oferta_interna_def']
    
    return None

def calcular_denominador_rapido(fecha_base, indexador, denominador, indexadores_completo_df):
    """
    Versión optimizada del cálculo de denominador usando DataFrame pre-cargado.
    MANTIENE la lógica exacta de la función original.
    """
    fecha_ano_mes = fecha_a_texto(fecha_base)
    
    # Buscar en el DataFrame combinado
    mask = indexadores_completo_df['fechaoperacion'].apply(lambda x: fecha_a_texto(x)) == fecha_ano_mes
    matching_rows = indexadores_completo_df[mask]
    
    if matching_rows.empty:
        return None
    
    row = matching_rows.iloc[0]
    
    # LÓGICA EXACTA de la función original
    if indexador == "IPC":
        return row['ipc']
    elif indexador != "IPC" and denominador == "PROVISIONAL":
        return row['oferta_interna_prov']
    elif indexador != "IPC" and denominador == "DEFINITIVO":
        return row['oferta_interna_def']
    
    return None

# ============================================================================
# CORREGIR EVALUACIÓN - USAR LÓGICA ORIGINAL EXACTA
# ============================================================================

def evaluar_oferta_corregido(precio_indexado, precio_sicep, precio_bolsa, constante_sicep=None, precio_fncer=None, es_oferta_fncer=False):
    """
    Evaluación de ofertas CORREGIDA - Usa la lógica exacta de la función original.
    """
    evaluacion = 0
    
    # Verificar que precio indexado sea válido (IGUAL que original)
    if precio_indexado is None or pd.isna(precio_indexado):
        return evaluacion
    
    # Evaluación para ofertas FNCER (LÓGICA EXACTA)
    if es_oferta_fncer:
        if precio_fncer is not None and precio_fncer > 0:
            if precio_indexado <= precio_fncer:
                evaluacion = 1
        else:
            # Fallback a evaluación normal (IGUAL que original)
            if precio_sicep is not None and precio_sicep > 0 and precio_bolsa is not None and precio_bolsa > 0:
                if constante_sicep is None:
                    constante_sicep = 1.0
                
                precio_sicep_ajustado = precio_sicep * constante_sicep
                limite = min(precio_sicep_ajustado, precio_bolsa)
                
                if precio_indexado <= limite:
                    evaluacion = 1
    else:
        # Evaluación para ofertas normales (LÓGICA EXACTA)
        if precio_sicep is not None and precio_bolsa is not None:
            if precio_sicep == 0 and precio_bolsa == 0:
                return evaluacion
            
            if precio_sicep > 0 or precio_bolsa > 0:
                if constante_sicep is None:
                    constante_sicep = 1.0
                
                precio_sicep_ajustado = precio_sicep * constante_sicep
                
                if precio_sicep_ajustado == 0 and precio_bolsa > 0:
                    limite = precio_bolsa
                elif precio_bolsa == 0 and precio_sicep_ajustado > 0:
                    limite = precio_sicep_ajustado
                else:
                    limite = min(precio_sicep_ajustado, precio_bolsa)
                
                if precio_indexado <= limite:
                    evaluacion = 1
    
    return evaluacion

# ============================================================================
# MANTENER FUNCIONES DE NORMALIZACIÓN (ESTAS ESTÁN BIEN)
# ============================================================================

def normalizar_columnas_precios_rapido(precios_df, codigo_oferta=""):
    """Versión optimizada de normalización de columnas de precios."""
    logger.debug(f"Normalizando precios para: {codigo_oferta}")
    
    nuevas_columnas = []
    
    patterns_compiled = [
        (re.compile(r'^\$/KWh-H(\d+)$', re.IGNORECASE), "$/KWh-H format"),
        (re.compile(r'^KWH-H(\d+)$', re.IGNORECASE), "KWH-H format"),
        (re.compile(r'^\$/KW[Hh]?[-_]H(\d+)$', re.IGNORECASE), "$/KW-H format"),
        (re.compile(r'^PRECIO[-_\s]*H(\d+)$', re.IGNORECASE), "PRECIO-H format"),
        (re.compile(r'^H(\d+)$', re.IGNORECASE), "H format"),
        (re.compile(r'^(\d+)$'), "number format"),
        (re.compile(r'.*[-_/]H(\d+)$', re.IGNORECASE), "generic-H format"),
        (re.compile(r'.*H(\d+)$', re.IGNORECASE), "ending-H format"),
    ]
    
    for col in precios_df.columns:
        col_str = str(col).strip()
        
        if col_str.upper() in ["FECHA", "DATE", "FECHA_OPERACION"]:
            nuevas_columnas.append("FECHA")
            continue
        
        hora_encontrada = None
        for pattern, descripcion in patterns_compiled:
            match = pattern.search(col_str)
            if match:
                try:
                    hora_encontrada = int(match.group(1))
                    if 1 <= hora_encontrada <= 24:
                        break
                except (ValueError, IndexError):
                    continue
        
        if hora_encontrada and 1 <= hora_encontrada <= 24:
            nuevas_columnas.append(f"H{hora_encontrada}")
        else:
            nuevas_columnas.append(col_str)
    
    precios_df_copy = precios_df.copy()
    precios_df_copy.columns = nuevas_columnas
    
    horas_encontradas = len([col for col in nuevas_columnas if col.startswith('H')])
    print(f"✅ {codigo_oferta} - Precios: {horas_encontradas}/24 columnas normalizadas")
    
    return precios_df_copy

def normalizar_columnas_cantidad_rapido(cantidad_df, codigo_oferta=""):
    """Versión optimizada de normalización de columnas de cantidad."""
    logger.debug(f"Normalizando cantidad para: {codigo_oferta}")
    
    nuevas_columnas = []
    
    patterns_compiled = [
        (re.compile(r'^KWH-H(\d+)$', re.IGNORECASE), "KWH-H format"),
        (re.compile(r'^KW[Hh][-_]H(\d+)$', re.IGNORECASE), "KWH_H format"),
        (re.compile(r'^CANTIDAD[-_\s]*H(\d+)$', re.IGNORECASE), "CANTIDAD-H format"),
        (re.compile(r'^H(\d+)$', re.IGNORECASE), "H format"),
        (re.compile(r'^(\d+)$'), "number format"),
        (re.compile(r'.*[-_/]H(\d+)$', re.IGNORECASE), "generic-H format"),
    ]
    
    for col in cantidad_df.columns:
        col_str = str(col).strip()
        
        if col_str.upper() in ["FECHA", "DATE", "FECHA_OPERACION"]:
            nuevas_columnas.append("FECHA")
            continue
        
        hora_encontrada = None
        for pattern, descripcion in patterns_compiled:
            match = pattern.search(col_str)
            if match:
                try:
                    hora_encontrada = int(match.group(1))
                    if 1 <= hora_encontrada <= 24:
                        break
                except (ValueError, IndexError):
                    continue
        
        if hora_encontrada and 1 <= hora_encontrada <= 24:
            nuevas_columnas.append(f"KWH-H{hora_encontrada}")
        else:
            nuevas_columnas.append(col_str)
    
    cantidad_df_copy = cantidad_df.copy()
    cantidad_df_copy.columns = nuevas_columnas
    
    horas_encontradas = len([col for col in nuevas_columnas if col.startswith('KWH-H')])
    print(f"✅ {codigo_oferta} - Cantidad: {horas_encontradas}/24 columnas normalizadas")
    
    return cantidad_df_copy

# ============================================================================
# FUNCIONES DE PRECIO OPTIMIZADAS (MANTENER ESTAS)
# ============================================================================

def procesar_precio_sicep_rapido(datos_iniciales=DATOS_INICIALES):
    """Versión optimizada del procesamiento de PRECIO SICEP."""
    logger.info(f"Procesando PRECIO SICEP (optimizado) desde {datos_iniciales}")
    
    if not verificar_archivo_existe(datos_iniciales):
        logger.error(f"No se encontró el archivo de datos iniciales: {datos_iniciales}")
        return None
    
    if not verificar_hoja_existe(datos_iniciales, "PROYECCIÓN PRECIO SICEP"):
        logger.info("No se encontró la hoja PROYECCIÓN PRECIO SICEP, se creará...")
        if not verificar_hoja_existe(datos_iniciales, "PRECIO SICEP"):
            logger.error("No se encontró la hoja PRECIO SICEP con los precios anuales")
            return None
        if not crear_proyeccion_precio_sicep(datos_iniciales):
            logger.error("No se pudo crear la proyección de precios SICEP")
            return None
    
    try:
        sicep_df = leer_excel_seguro(datos_iniciales, "PROYECCIÓN PRECIO SICEP")
        if sicep_df.empty:
            logger.error("La hoja PROYECCIÓN PRECIO SICEP está vacía")
            return None
        
        sicep_df['FECHA'] = pd.to_datetime(sicep_df['FECHA'], errors='coerce').dt.date
        sicep_df = sicep_df.dropna(subset=['FECHA'])
        
        if 'PRECIO' not in sicep_df.columns:
            logger.error("No se encontró la columna 'PRECIO' en la hoja PROYECCIÓN PRECIO SICEP")
            return None
            
        sicep_df['AUX'] = sicep_df['FECHA'].apply(lambda d: f"{d.year}-{d.month}")
        sicep_dict = {row['AUX']: row['PRECIO'] for _, row in sicep_df.iterrows()}
        
        fncer_dict = {}
        if 'PRECIO FNCER' in sicep_df.columns:
            fncer_dict = {row['AUX']: row['PRECIO FNCER'] for _, row in sicep_df.iterrows()}
            print(f"PRECIO SICEP y FNCER procesados (optimizado): {len(sicep_dict)} períodos")
        else:
            print(f"PRECIO SICEP procesado (optimizado): {len(sicep_dict)} períodos")
        
        return {'SICEP': sicep_dict, 'FNCER': fncer_dict}
    
    except Exception as e:
        logger.exception(f"Error al procesar PROYECCIÓN PRECIO SICEP: {e}")
        return None

def procesar_precio_bolsa_rapido(datos_iniciales=DATOS_INICIALES):
    """Versión optimizada del procesamiento de PRECIO BOLSA."""
    logger.info(f"Procesando PRECIO BOLSA (optimizado) desde {datos_iniciales}")
    
    if not verificar_archivo_existe(datos_iniciales):
        return None
    
    if not verificar_hoja_existe(datos_iniciales, "P BOLSA"):
        logger.error(f"No se encontró la hoja 'P BOLSA' en el archivo: {datos_iniciales}")
        return None
    
    try:
        bolsa_df = leer_excel_seguro(datos_iniciales, "P BOLSA")
        if bolsa_df.empty:
            return None
        
        bolsa_df['FECHA'] = pd.to_datetime(bolsa_df['FECHA'], format="%d/%m/%Y", errors='coerce').dt.date
        bolsa_df = bolsa_df.dropna(subset=['FECHA'])
        bolsa_df['AUX'] = bolsa_df['FECHA'].apply(lambda d: f"{d.year}-{d.month}")
        bolsa_dict = {row['AUX']: row['PBNA'] for _, row in bolsa_df.iterrows()}
        
        print(f"PRECIO BOLSA procesado (optimizado): {len(bolsa_dict)} períodos")
        return bolsa_dict
    
    except Exception as e:
        logger.exception(f"Error al procesar PRECIO BOLSA: {e}")
        return None

# ============================================================================
# FUNCIÓN PRINCIPAL CORREGIDA
# ============================================================================

def procesar_ofertas_optimizado_corregido(carpeta_ofertas=OFERTAS_DIR, datos_iniciales=DATOS_INICIALES, 
                                        archivo_salida=RESULTADO_OFERTAS):
    """
    Versión OPTIMIZADA CORREGIDA con VALIDACIÓN DE NOMBRES ESTANDARIZADOS.
    MANTIENE la velocidad pero AGREGA validación automática de nombres.
    Formato esperado: Agente-OFERTA-# (ejemplo: EPM-OFERTA-001.xlsx)
    """
    print("🚀 INICIANDO PROCESAMIENTO OPTIMIZADO CON ESTANDARIZACIÓN")
    start_total = time.time()
    
    logger.info(f"Procesando ofertas (OPTIMIZADO CON ESTANDARIZACIÓN) en {carpeta_ofertas}")
    
    # Verificaciones iniciales (sin cambios)
    if not verificar_archivo_existe(datos_iniciales):
        logger.error(f"No se encontró el archivo de datos iniciales: {datos_iniciales}")
        return False
    
    if not os.path.exists(carpeta_ofertas):
        logger.error(f"No se encontró la carpeta de ofertas: {carpeta_ofertas}")
        return False
    
    # Solicitar constante SICEP (sin cambios)
    try:
        constante_sicep = solicitar_input_seguro(
            "Ingrese la constante para el cálculo del precio SICEP: ",
            tipo=float,
            validacion=lambda x: x > 0,
            mensaje_error="La constante debe ser un número positivo."
        )
        print(f"Usando constante SICEP: {constante_sicep}")
    except Exception as e:
        logger.warning(f"Error al solicitar constante SICEP: {e}. Usando valor predeterminado.")
        constante_sicep = 1.0
        print(f"Usando constante SICEP predeterminada: {constante_sicep}")
    
    # =========================================================================
    # FASE 1: PRE-CARGA OPTIMIZADA (sin cambios)
    # =========================================================================
    
    print("\n📊 FASE 1: Pre-cargando datos compartidos...")
    start_fase1 = time.time()
    
    # Leer indexadores con case-insensitive
    print("🔍 Buscando hoja INDEXADORES...")
    existe_indexadores, nombre_real_indexadores = verificar_hoja_existe_case_insensitive(datos_iniciales, "INDEXADORES")
    if not existe_indexadores:
        logger.error("No se encontró la hoja INDEXADORES")
        return False
    else:
        if nombre_real_indexadores != "INDEXADORES":
            print(f"✅ Encontrada: 'INDEXADORES' → '{nombre_real_indexadores}'")
    
    indexadores_df = leer_excel_case_insensitive(datos_iniciales, "INDEXADORES")
    if indexadores_df.empty:
        logger.error("La hoja INDEXADORES está vacía")
        return False
    
    # Leer proyección con case-insensitive
    print("🔍 Buscando hoja PROYECCIÓN INDEXADORES...")
    existe_proyeccion, nombre_real_proyeccion = verificar_hoja_existe_case_insensitive(datos_iniciales, "PROYECCIÓN INDEXADORES")
    if not existe_proyeccion:
        logger.warning("No se encontró PROYECCIÓN INDEXADORES, se creará automáticamente")
        from core.indexadores import crear_proyeccion_indexadores
        if not crear_proyeccion_indexadores(datos_iniciales, carpeta_ofertas):
            logger.error("No se pudo crear la proyección de indexadores")
            return False
        proyeccion_df = leer_excel_case_insensitive(datos_iniciales, "PROYECCIÓN INDEXADORES")
    else:
        if nombre_real_proyeccion != "PROYECCIÓN INDEXADORES":
            print(f"✅ Encontrada: 'PROYECCIÓN INDEXADORES' → '{nombre_real_proyeccion}'")
        proyeccion_df = leer_excel_case_insensitive(datos_iniciales, "PROYECCIÓN INDEXADORES")
    
    if proyeccion_df.empty:
        logger.error("La proyección de indexadores está vacía")
        return False
    
    # Convertir fechas UNA sola vez (sin cambios)
    indexadores_df['fechaoperacion'] = pd.to_datetime(indexadores_df['fechaoperacion'], format="%d/%m/%Y").dt.date
    proyeccion_df['fechaoperacion'] = pd.to_datetime(proyeccion_df['fechaoperacion'], format="%d/%m/%Y").dt.date
    
    # Pre-computar indexadores corregidos (sin cambios)
    indexadores_completo_df = precomputar_indexadores_corregido(indexadores_df, proyeccion_df)
    
    # Procesar precios SICEP y BOLSA (sin cambios)
    sicep_dict = procesar_precio_sicep_rapido(datos_iniciales)
    if sicep_dict is None:
        logger.error("No se pudo procesar PRECIO SICEP")
        return False
    
    bolsa_dict = procesar_precio_bolsa_rapido(datos_iniciales)
    if bolsa_dict is None:
        logger.error("No se pudo procesar PRECIO BOLSA")
        return False
    
    end_fase1 = time.time()
    print(f"✅ FASE 1 completada en {end_fase1 - start_fase1:.1f} segundos")
    
    # =========================================================================
    # FASE 2: PROCESAMIENTO OPTIMIZADO CON VALIDACIÓN DE NOMBRES
    # =========================================================================
    
    print("\n🔄 FASE 2: Procesamiento optimizado con validación de nombres...")
    start_fase2 = time.time()
    
    # Buscar archivos (sin cambios)
    archivos = [f for f in os.listdir(carpeta_ofertas) 
               if f.endswith('.xlsx') and not f.startswith('~$')]
    
    if not archivos:
        logger.error(f"No se encontraron archivos en {carpeta_ofertas}")
        return False
    
    # ===== NUEVA FUNCIONALIDAD: VALIDACIÓN DE NOMBRES =====
    print(f"\n📋 Validando nombres de {len(archivos)} ofertas...")
    
    from core.validador_ofertas import ValidadorNombresOfertas
    validador = ValidadorNombresOfertas()
    
    # Validar cada archivo
    ofertas_validas = []
    ofertas_invalidas = []
    
    for archivo in archivos:
        oferta_info = validador.validar_nombre_archivo(archivo)
        if oferta_info.es_valido:
            ofertas_validas.append(oferta_info)
        else:
            ofertas_invalidas.append(oferta_info)
    
    # Mostrar resumen de validación
    resumen = validador.obtener_resumen_validacion()
    print(f"📊 VALIDACIÓN: {resumen['ofertas_validas']} válidas, {resumen['ofertas_invalidas']} inválidas")
    
    if resumen['agentes_encontrados']:
        print(f"🏢 Agentes: {', '.join(resumen['agentes_encontrados'])}")
    
    if ofertas_invalidas:
        print(f"\n⚠️ OFERTAS CON NOMBRES INVÁLIDOS:")
        for oferta in ofertas_invalidas:
            print(f"   ❌ {oferta.nombre_archivo}: {oferta.error_mensaje}")
        
        print(f"\n💡 FORMATO CORRECTO: Agente-OFERTA-Número")
        print(f"   Ejemplos: EPM-OFERTA-001.xlsx, AES-OFERTA-002.xlsx")
        
        # Preguntar si continuar
        if resumen['ofertas_validas'] > 0:
            continuar = input(f"\n¿Continuar con las {resumen['ofertas_validas']} ofertas válidas? (s/n): ")
            if continuar.lower() != 's':
                print("❌ Procesamiento cancelado")
                return False
        else:
            print("❌ No hay ofertas válidas para procesar")
            return False
    
    print(f"🚀 Procesando {resumen['ofertas_validas']} ofertas válidas con lectura optimizada...\n")
    # ===== FIN DE VALIDACIÓN =====
    
    # Inicializar contenedores de resultados (sin cambios)
    tabla_maestra = []
    cantidades_precios = []
    
    # Contadores de rendimiento (sin cambios)
    ofertas_exitosas = 0
    ofertas_con_errores = 0
    total_registros = 0
    
    # ===== PROCESAR SOLO LAS OFERTAS VÁLIDAS =====
    for i, oferta_info in enumerate(ofertas_validas, 1):
        # USAR INFORMACIÓN DE LA VALIDACIÓN
        codigo_oferta = oferta_info.nombre_estandarizado  # Nombre estandarizado
        archivo = oferta_info.nombre_archivo             # Archivo original
        agente = oferta_info.agente                      # Agente extraído
        numero = oferta_info.numero                      # Número extraído
        ruta_archivo = os.path.join(carpeta_ofertas, archivo)
        
        print(f"\n🔄 [{i}/{len(ofertas_validas)}] Procesando: {codigo_oferta}")
        print(f"   📁 Archivo: {archivo}")
        print(f"   🏢 Agente: {agente}")
        print(f"   #️⃣ Número: {numero}")
        start_oferta = time.time()
        
        try:
            # LECTURA OPTIMIZADA (sin cambios)
            print(f"📖 Leyendo hojas (optimizada): INDEXADOR, cantidad, precios")
            hojas_data = leer_excel_ultra_rapido(ruta_archivo, ["INDEXADOR", "cantidad", "precios"])
            
            indexador_df = hojas_data.get("INDEXADOR", pd.DataFrame())
            cantidad_df = hojas_data.get("cantidad", pd.DataFrame())
            precios_df = hojas_data.get("precios", pd.DataFrame())
            
            if indexador_df.empty or cantidad_df.empty or precios_df.empty:
                logger.error(f"Error al leer hojas de {archivo}")
                ofertas_con_errores += 1
                print(f"❌ {codigo_oferta}: Error en lectura de hojas")
                continue
            
            # Normalización optimizada (sin cambios)
            precios_df = normalizar_columnas_precios_rapido(precios_df, codigo_oferta)
            cantidad_df = normalizar_columnas_cantidad_rapido(cantidad_df, codigo_oferta)
            
            # Verificación de columnas (sin cambios)
            precios_cols = [col for col in precios_df.columns if col.startswith('H')]
            cantidad_cols = [col for col in cantidad_df.columns if col.startswith('KWH-H')]
            
            if len(precios_cols) == 0 or len(cantidad_cols) == 0:
                ofertas_con_errores += 1
                print(f"❌ {codigo_oferta}: Columnas insuficientes")
                continue
            
            # Conversión de fechas (sin cambios)
            cantidad_df['FECHA'] = pd.to_datetime(cantidad_df['FECHA'], format="%d/%m/%Y", errors='coerce').dt.date
            precios_df['FECHA'] = pd.to_datetime(precios_df['FECHA'], format="%d/%m/%Y", errors='coerce').dt.date
            
            cantidad_df = cantidad_df.dropna(subset=['FECHA'])
            precios_df = precios_df.dropna(subset=['FECHA'])
            
            # ===== EXTRACCIÓN DE METADATA CON CAMPOS NUEVOS =====
            try:
                indexador_data = {
                    "CÓDIGO OFERTA": codigo_oferta,      # Nombre estandarizado
                    "AGENTE": agente,                    # NUEVO: Agente extraído
                    "NUMERO": numero,                    # NUEVO: Número extraído
                    "ARCHIVO ORIGINAL": archivo,         # NUEVO: Nombre archivo original
                    "INDEXADOR": indexador_df.loc[indexador_df["CONCEPTO"] == "INDEXADOR", "VALOR"].values[0],
                    "NUMERADOR": indexador_df.loc[indexador_df["CONCEPTO"] == "NUMERADOR", "VALOR"].values[0],
                    "DENOMINADOR": indexador_df.loc[indexador_df["CONCEPTO"] == "DENOMINADOR", "VALOR"].values[0],
                    "FECHA BASE": pd.to_datetime(indexador_df.loc[indexador_df["CONCEPTO"] == "FECHA BASE", "VALOR"].values[0]).date()
                }
                
                # Verificar FNCER (sin cambios)
                fncer_rows = indexador_df[indexador_df["CONCEPTO"] == "FNCER"]
                if not fncer_rows.empty:
                    indexador_data["FNCER"] = "SI" if fncer_rows["VALOR"].values[0].upper() == "SI" else "NO"
                else:
                    indexador_data["FNCER"] = "NO"
                
                tabla_maestra.append(indexador_data)
                
            except Exception as e:
                logger.error(f"Error en metadata de {codigo_oferta}: {e}")
                ofertas_con_errores += 1
                continue
            
            # Determinar si es FNCER (sin cambios)
            es_fncer = indexador_data.get("FNCER", "NO") == "SI"
            registros_oferta = 0
            
            # ===== PROCESAMIENTO OPTIMIZADO DE DATOS (sin cambios en lógica) =====
            for fecha in cantidad_df['FECHA'].unique():
                if pd.isna(fecha):
                    continue
                
                # Obtener filas para esta fecha
                cantidad_row = cantidad_df[cantidad_df['FECHA'] == fecha]
                precio_row = precios_df[precios_df['FECHA'] == fecha]
                
                if cantidad_row.empty or precio_row.empty:
                    continue
                
                cantidad_row = cantidad_row.iloc[0]
                precio_row = precio_row.iloc[0]
                
                # Precios para esta fecha
                fecha_aux = f"{fecha.year}-{fecha.month}"
                precio_sicep_val = sicep_dict.get('SICEP', {}).get(fecha_aux, 0)
                precio_bolsa_val = bolsa_dict.get(fecha_aux, 0)
                precio_fncer_val = sicep_dict.get('FNCER', {}).get(fecha_aux, 0) if es_fncer else None
                
                # Procesar las 24 horas para esta fecha
                for hora in range(1, 25):
                    cantidad = cantidad_row.get(f"KWH-H{hora}", 0)
                    precio = precio_row.get(f"H{hora}", None)
                    
                    if cantidad > 0 and precio is not None and pd.notna(precio):
                        # CÁLCULO CORREGIDO DE INDEXACIÓN (sin cambios)
                        numerador_valor = calcular_numerador_rapido(
                            fecha,
                            indexador_data["INDEXADOR"],
                            indexador_data["NUMERADOR"],
                            indexadores_completo_df
                        )
                        
                        denominador_valor = calcular_denominador_rapido(
                            indexador_data["FECHA BASE"],
                            indexador_data["INDEXADOR"],
                            indexador_data["DENOMINADOR"],
                            indexadores_completo_df
                        )
                        
                        # Calcular precio indexado (sin cambios)
                        if (
                            precio is not None
                            and numerador_valor is not None
                            and denominador_valor is not None
                            and denominador_valor != 0
                        ):
                            precio_indexado = (float(precio) + 0) * ((float(numerador_valor) + 0) / (float(denominador_valor) + 0))
                        else:
                            precio_indexado = None
                        
                        # EVALUACIÓN CORREGIDA (sin cambios)
                        evaluacion = evaluar_oferta_corregido(
                            precio_indexado,
                            precio_sicep_val,
                            precio_bolsa_val,
                            constante_sicep,
                            precio_fncer=precio_fncer_val,
                            es_oferta_fncer=es_fncer
                        )
                        
                        # ===== CREAR REGISTRO CON CAMPOS NUEVOS =====
                        registro = {
                            "CÓDIGO OFERTA": codigo_oferta,         # Nombre estandarizado
                            "AGENTE": agente,                       # NUEVO: Agente
                            "NUMERO": numero,                       # NUEVO: Número  
                            "ARCHIVO ORIGINAL": archivo,            # NUEVO: Archivo original
                            "FECHA": fecha,
                            "Atributo": hora,
                            "CANTIDAD": float(cantidad),
                            "PRECIO": float(precio),
                            "INDEXADOR": indexador_data["INDEXADOR"],
                            "NUMERADOR": indexador_data["NUMERADOR"],
                            "DENOMINADOR": indexador_data["DENOMINADOR"],
                            "FECHA BASE": indexador_data["FECHA BASE"],
                            "NUMERADOR #": numerador_valor,
                            "DENOMINADOR #": denominador_valor,
                            "PRECIO INDEXADO": precio_indexado,
                            "FNCER": indexador_data.get("FNCER", "NO"),
                            "PRECIO SICEP": precio_fncer_val if es_fncer and precio_fncer_val else precio_sicep_val,
                            "PRECIO BOLSA": precio_bolsa_val,
                            "EVALUACIÓN": evaluacion
                        }
                        
                        cantidades_precios.append(registro)
                        registros_oferta += 1
            
            # Estadísticas de la oferta (mejoradas)
            end_oferta = time.time()
            tiempo_oferta = end_oferta - start_oferta
            
            if registros_oferta > 0:
                ofertas_exitosas += 1
                total_registros += registros_oferta
                print(f"✅ {codigo_oferta} ({agente}): {registros_oferta:,} registros en {tiempo_oferta:.1f}s")
            else:
                ofertas_con_errores += 1
                print(f"❌ {codigo_oferta} ({agente}): 0 registros generados")
        
        except Exception as e:
            logger.error(f"Error procesando {codigo_oferta}: {e}")
            ofertas_con_errores += 1
            print(f"❌ {codigo_oferta} ({agente}): Error - {e}")
            continue
    
    end_fase2 = time.time()
    print(f"\n✅ FASE 2 completada en {end_fase2 - start_fase2:.1f} segundos")
    
    # =========================================================================
    # FASE 3: GUARDAR RESULTADOS (sin cambios)
    # =========================================================================
    
    print("\n💾 FASE 3: Guardando resultados...")
    start_fase3 = time.time()
    
    # Verificar que tengamos datos
    if not tabla_maestra or not cantidades_precios:
        logger.error("No se generaron datos para guardar")
        print(f"❌ ERROR: No hay datos para guardar")
        return False
    
    try:
        # Crear DataFrames
        tabla_maestra_df = pd.DataFrame(tabla_maestra)
        cantidades_precios_df = pd.DataFrame(cantidades_precios)
        
        # Crear directorios si no existen
        Path(archivo_salida).parent.mkdir(parents=True, exist_ok=True)
        
        # Guardar con engine optimizado
        with pd.ExcelWriter(archivo_salida, engine="openpyxl") as writer:
            tabla_maestra_df.to_excel(writer, sheet_name="TABLA MAESTRA OFERTAS", index=False)
            cantidades_precios_df.to_excel(writer, sheet_name="CANTIDADES Y PRECIOS", index=False)
        
        end_fase3 = time.time()
        print(f"✅ FASE 3 completada en {end_fase3 - start_fase3:.1f} segundos")
        
        # ===== ESTADÍSTICAS FINALES MEJORADAS =====
        end_total = time.time()
        tiempo_total = end_total - start_total
        
        print(f"\n🎉 PROCESAMIENTO OPTIMIZADO CON ESTANDARIZACIÓN COMPLETADO:")
        print(f"   ⚡ Tiempo total: {tiempo_total:.1f} segundos")
        print(f"   📊 Ofertas exitosas: {ofertas_exitosas}/{len(ofertas_validas)}")
        print(f"   ❌ Ofertas con errores: {ofertas_con_errores}")
        print(f"   📈 Total registros: {total_registros:,}")
        print(f"   ⚡ Velocidad: {total_registros/tiempo_total:.0f} registros/segundo")
        print(f"   🏢 Agentes procesados: {', '.join(sorted(resumen['agentes_encontrados']))}")
        print(f"   💾 Archivo guardado: {archivo_salida}")
        print(f"   ✅ ESTANDARIZACIÓN: Nombres validados automáticamente")
        
        # Desglose por agente MEJORADO
        if cantidades_precios_df is not None and not cantidades_precios_df.empty:
            print(f"\n📈 DESGLOSE POR AGENTE:")
            agente_stats = cantidades_precios_df.groupby('AGENTE').agg({
                'CÓDIGO OFERTA': 'nunique',
                'EVALUACIÓN': ['count', 'sum']
            }).round(2)
            
            for agente in agente_stats.index:
                ofertas_agente = agente_stats.loc[agente, ('CÓDIGO OFERTA', 'nunique')]
                total_registros_agente = agente_stats.loc[agente, ('EVALUACIÓN', 'count')]
                aprobados = agente_stats.loc[agente, ('EVALUACIÓN', 'sum')]
                porcentaje_aprobacion = (aprobados / total_registros_agente * 100) if total_registros_agente > 0 else 0
                
                print(f"   🏢 {agente}: {ofertas_agente} ofertas, {total_registros_agente:,} registros, {porcentaje_aprobacion:.1f}% aprobados")
            
            # Evaluaciones globales
            evaluaciones = cantidades_precios_df['EVALUACIÓN'].value_counts()
            print(f"\n📈 RESULTADOS DE EVALUACIÓN GLOBAL:")
            for eval_val, count in evaluaciones.items():
                status = "✅ APROBADOS" if eval_val == 1 else "❌ RECHAZADOS"
                percentage = (count / len(cantidades_precios_df)) * 100
                print(f"   {status}: {count:,} registros ({percentage:.1f}%)")
        
        logger.info(f"Procesamiento optimizado con estandarización completado en {tiempo_total:.1f}s")
        return True
        
    except Exception as e:
        logger.error(f"Error al guardar resultados: {e}")
        print(f"❌ ERROR al guardar: {e}")
        return False
    